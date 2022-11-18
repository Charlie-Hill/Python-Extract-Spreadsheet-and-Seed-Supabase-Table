from typing import OrderedDict
import openpyxl
from pathlib import Path
import json

headings = []
headingsRow = 8 # TODO: Extract this to config or env file?
instrumentsStartRow = 9  # TODO: Automatically find first row
instrumentsEndRow = 4532 # TODO: Automatically find last row

def main():
    # Load xlsx file
    print("Loading Instrument list_30.xlsx...")
    xlsx_file = Path('./data/lse/Instrument list_30.xlsx')
    wb_obj = openpyxl.load_workbook(xlsx_file)

    # Set active sheet to 1.0 All Equity
    print("Setting active sheet...")
    wb_obj.active = 0
    sheet = wb_obj.active

    # Load column headings
    print("Loading column headings...")
    columnHeadingsList = columnHeadings(sheet)
    headingsTxt = Path('headings.txt')
    with open(headingsTxt, 'w') as f:
        for item in columnHeadingsList:
            f.write("%s\n" % item)
    print("Loaded column headings.")

    # Load row values
    print("Loading instruments...")
    instruments = []
    for i in range(instrumentsStartRow, instrumentsEndRow):
        instruments.append(rowValues(sheet, i))
    with open('instruments.json', 'w') as f:
        f.write(json.dumps(instruments, indent=4, default=str))
    print("Loaded instruments.")

    print("Finished!")

def columnHeadings(sheet):
    for i in range(1, sheet.max_column + 1):
        cell_obj = sheet.cell(row = headingsRow, column = i)
        heading = cell_obj.value.replace(' ', '_').replace('-', '_').lower()
        if heading == "tidm":
            heading = heading.upper()
        headings.append(heading)
    return headings

def rowValues(sheet, rowNum):
    instrument = OrderedDict()
    
    for i in range(1, sheet.max_column + 1):
        cell_obj = sheet.cell(row = rowNum, column = i)
        instrument[headings[i - 1]] = cell_obj.value
        
    return instrument

main()